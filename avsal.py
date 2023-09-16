#!/usr/bin/env python
# coding: utf-8

# In[1]:

import pandas as pd

# In[3]:

import openpyxl

import numpy as np
import scipy.stats as st
from scipy.stats import t

# File directive
path = "python_moscow.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

# In[4]:

row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total number of rows:", row)
print("Total number of columns:", column)

# In[50]:

avsal = []  # Average salaries for each level
error = []  # Standard deviation of average salaries
vol = []  # Sample sizes
alpha = 0.1  # Confidence coefficient: 1-alpha

avsalary = []
level = []
exp = []
experience = []
minsalary = []
maxsalary = []

for i in range(2, row+1):
    cell_obj = sheet_obj.cell(i, 6)
    minsalary.append(cell_obj.value)
for i in range(2, row+1):
    cell_obj = sheet_obj.cell(i, 7)
    maxsalary.append(cell_obj.value)
for i in range(2, row+1):
    cell_obj = sheet_obj.cell(i, 2)
    level.append(cell_obj.value)
for i in range(2, row+1):
    cell_obj = sheet_obj.cell(i, 4)
    experience.append(cell_obj.value)

for j in range(0, 4):  # From the minimum to the maximum specialist level

    i = 0

    xo = 0  # Average salary for the level

    # Calculate the average value
    i = 0
    while (i < row-1):  # Select positions with specified salaries
        if (level[i] != j+1) or (minsalary[i] == 0 and maxsalary[i] == 0):
            i = i+1
        else:
            if minsalary[i] == 0 or maxsalary[i] == 0:  # If only min/max salary is specified, take it as the salary value
                avsalary.append(minsalary[i] + maxsalary[i])
            else:  # If a range is specified, take the arithmetic mean as the salary
                avsalary.append((minsalary[i] + maxsalary[i])/2)

            exp.append(experience[i])

            i = i+1

    n = len(avsalary)
    xo = sum(avsalary) / len(avsalary)

    # Exclude outliers
    outliersal = []
    outliersexp = []

    std = np.std(avsalary, ddof=1)  # Sample standard deviation
    stud = t.ppf(1 - alpha/2, len(avsalary)-1)  # Student's coefficient for the given confidence interval

    i = 0

    while (i != n):  # Values that significantly deviate at the alpha% significance level are identified and excluded
        if (abs(avsalary[i]-xo)/std > stud):
            outlierexp = exp.pop(i)
            outliersexp.append(outlierexp)
            outlier = avsalary.pop(i)
            outliersal.append(outlier)
            n = len(avsalary)
            i = i+1
        else:
            i = i+1

    print("\nOutliers for level", j+1, ":", outliersal)
    print("\nExperience for level", j+1, ":", outliersexp)

    # Calculate the average and error
    avsal.append(sum(avsalary) / len(avsalary))
    error.append(st.sem(avsalary))
    vol.append(len(avsalary))

print("\nAverage salaries:", avsal)
print("\nRandom error:", error)
print("\nSample size:", vol)

# In[51]:

import matplotlib.pyplot as plt

labels = ['intern', 'junior', 'middle', 'senior']
x_pos = np.arange(len(labels))

fig, ax = plt.subplots()
ax.bar(x_pos, avsal,
       yerr=error,
       align='center',
       alpha=0.5,
       ecolor='black',
       capsize=10)
ax.set_ylabel('salary, rubles')
ax.set_xticks(x_pos)
ax.set_xticklabels(labels)
ax.set_title('Average salaries')

plt.show()


# In[ ]:




